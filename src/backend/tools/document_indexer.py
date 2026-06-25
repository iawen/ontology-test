"""
Schema Optimizer v3.1 - 文档分块与向量化模块
============================================
改进点：
  1. 降级解析链：fitz → pdfplumber → 纯文本兜底
  2. 分块重叠（chunk_overlap）：避免语义边界丢失
  3. 元数据增强：提取标题层级
  4. Embedding 缓存：避免重复计算
  5. 本地备选模型：API失败时降级到 sentence-transformers
"""

import os
import re
import json
import hashlib
import zipfile
from pathlib import Path
from xml.etree import ElementTree
from typing import Optional, List, Dict, Tuple, Callable
from functools import lru_cache

import numpy as np


# ============================================================
# 文档解析（降级链 + 重叠分块）
# ============================================================

def parse_document(path: Path, chunk_size: int = 800, chunk_overlap: int = 150) -> List[Dict]:
    """
    解析文档为结构化分块列表。

    Args:
        path: 文档路径
        chunk_size: 每块目标字符数
        chunk_overlap: 块间重叠字符数（防止边界信息丢失）

    Returns:
        [{"chunk_id": str, "content": str, "metadata": {"source": str, "page": int, "type": str, "section": str}}]
    """
    ext = path.suffix.lower()
    if ext == ".docx":
        return _parse_docx(path, chunk_size, chunk_overlap)
    elif ext == ".pdf":
        return _parse_pdf_with_fallback(path, chunk_size, chunk_overlap)
    elif ext == ".xlsx":
        return _parse_xlsx(path, chunk_size, chunk_overlap)
    elif ext == ".txt":
        return _parse_txt(path, chunk_size, chunk_overlap)
    else:
        return []


def _split_text_with_overlap(text: str, chunk_size: int, chunk_overlap: int, metadata: dict) -> List[Dict]:
    """按字符数切分文本，支持重叠"""
    if len(text) <= chunk_size:
        return [{
            "chunk_id": hashlib.md5(f"{metadata.get('source','')}:{text[:50]}".encode()).hexdigest()[:12],
            "content": text,
            "metadata": metadata.copy()
        }]

    chunks = []
    start = 0
    idx = 0
    while start < len(text):
        end = min(start + chunk_size, len(text))
        # 尽量在句号/换行处切断，避免截断词语
        if end < len(text):
            # 找最近的句号、问号、感叹号或换行
            for sep in ["。", "！", "？", "\n\n", "\n", ". ", "! ", "? "]:
                pos = text.rfind(sep, start, end)
                if pos != -1 and pos > start + chunk_size // 2:
                    end = pos + len(sep)
                    break
        chunk_text = text[start:end].strip()
        if chunk_text:
            chunk_id = hashlib.md5(f"{metadata.get('source','')}:{idx}:{chunk_text[:30]}".encode()).hexdigest()[:12]
            chunks.append({
                "chunk_id": chunk_id,
                "content": chunk_text,
                "metadata": metadata.copy()
            })
        start = end - chunk_overlap if end < len(text) else end
        idx += 1
    return chunks


def _parse_docx(path: Path, chunk_size: int, chunk_overlap: int) -> List[Dict]:
    """解析 Word 文档，提取段落 + 标题层级"""
    ns = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    all_chunks = []
    try:
        with zipfile.ZipFile(path) as archive:
            xml = archive.read("word/document.xml")
        root = ElementTree.fromstring(xml)

        current_section = ""
        section_level = 0
        paragraphs = []

        for para in root.findall(".//w:p", ns):
            texts = [t.text for t in para.findall(".//w:t", ns) if t.text]
            text = "".join(texts).strip()
            if not text:
                continue

            # 检测标题样式（Heading1, Heading2, ...）
            style_elems = para.findall(".//w:pStyle", ns)
            style = ""
            if style_elems:
                style = style_elems[0].get("{http://schemas.openxmlformats.org/wordprocessingml/2006/main}val") or ""

            if style and "Heading" in style:
                try:
                    section_level = int(style.replace("Heading", "")) if style.replace("Heading", "").isdigit() else 0
                except:
                    section_level = 0
                current_section = text
                # 标题本身也作为内容保留
                paragraphs.append((text, current_section, section_level))
            else:
                paragraphs.append((text, current_section, section_level))

        # 合并段落为文本块
        full_text = "\n".join([p[0] for p in paragraphs])
        metadata = {
            "source": path.name,
            "type": "docx",
            "section": current_section,
            "section_level": section_level,
        }
        all_chunks = _split_text_with_overlap(full_text, chunk_size, chunk_overlap, metadata)

    except Exception as e:
        print(f"[Warning] 解析 DOCX 失败: {e}")
        # 降级：尝试直接读文本
        try:
            text = path.read_text(encoding="utf-8", errors="ignore")
            metadata = {"source": path.name, "type": "docx_fallback"}
            all_chunks = _split_text_with_overlap(text, chunk_size, chunk_overlap, metadata)
        except:
            pass
    return all_chunks


def _parse_pdf_with_fallback(path: Path, chunk_size: int, chunk_overlap: int) -> List[Dict]:
    """解析 PDF，带降级链：fitz → pdfplumber → 纯文本兜底"""
    chunks = []

    # 尝试1: PyMuPDF (fitz)
    try:
        import fitz
        doc = fitz.open(str(path))
        full_text = ""
        for page_num in range(len(doc)):
            page = doc[page_num]
            text = page.get_text().strip()
            if text:
                full_text += f"\n--- Page {page_num+1} ---\n{text}"
        doc.close()
        if full_text.strip():
            metadata = {"source": path.name, "type": "pdf_fitz"}
            return _split_text_with_overlap(full_text, chunk_size, chunk_overlap, metadata)
    except ImportError:
        pass
    except Exception as e:
        print(f"[Warning] PyMuPDF 解析失败: {e}")

    # 尝试2: pdfplumber
    try:
        import pdfplumber
        full_text = ""
        with pdfplumber.open(str(path)) as pdf:
            for page_num, page in enumerate(pdf.pages):
                text = (page.extract_text() or "").strip()
                if text:
                    full_text += f"\n--- Page {page_num+1} ---\n{text}"
        if full_text.strip():
            metadata = {"source": path.name, "type": "pdf_plumber"}
            return _split_text_with_overlap(full_text, chunk_size, chunk_overlap, metadata)
    except ImportError:
        pass
    except Exception as e:
        print(f"[Warning] pdfplumber 解析失败: {e}")

    # 尝试3: 纯文本兜底（仅提取可读字符）
    try:
        raw = path.read_bytes()
        # 尝试常见编码
        for encoding in ["utf-8", "gb18030", "latin-1"]:
            try:
                text = raw.decode(encoding, errors="ignore")
                # 过滤掉不可打印字符
                text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", " ", text)
                text = re.sub(r"\s+", " ", text)
                if len(text.strip()) > 100:
                    metadata = {"source": path.name, "type": "pdf_raw_text"}
                    return _split_text_with_overlap(text, chunk_size, chunk_overlap, metadata)
            except:
                continue
    except Exception as e:
        print(f"[Warning] PDF 纯文本兜底失败: {e}")

    print(f"[Warning] 所有 PDF 解析方式均失败: {path.name}")
    return chunks


def _parse_xlsx(path: Path, chunk_size: int, chunk_overlap: int) -> List[Dict]:
    """解析 Excel，按 sheet 分块"""
    chunks = []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(str(path), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h else f"col_{j}" for j, h in enumerate(rows[0])]
            # 将整个 sheet 转为文本
            lines = []
            for row in rows[1:]:  # 跳过表头
                row_dict = {headers[j]: str(row[j]) if j < len(row) and row[j] else "" for j in range(len(headers))}
                lines.append(json.dumps(row_dict, ensure_ascii=False))
            full_text = f"Sheet: {sheet_name}\nHeaders: {', '.join(headers)}\n" + "\n".join(lines)
            metadata = {"source": path.name, "sheet": sheet_name, "type": "excel"}
            chunks.extend(_split_text_with_overlap(full_text, chunk_size, chunk_overlap, metadata))
        wb.close()
    except Exception as e:
        print(f"[Warning] 解析 XLSX 失败: {e}")
    return chunks


def _parse_txt(path: Path, chunk_size: int, chunk_overlap: int) -> List[Dict]:
    """解析纯文本"""
    try:
        text = path.read_text(encoding="utf-8")
        metadata = {"source": path.name, "type": "text"}
        return _split_text_with_overlap(text, chunk_size, chunk_overlap, metadata)
    except Exception as e:
        print(f"[Warning] 解析 TXT 失败: {e}")
        return []


# ============================================================
# 向量化与语义检索（带缓存 + 本地备选）
# ============================================================

class DocumentIndex:
    """文档向量索引（内存级，带缓存）"""

    def __init__(self, embedding_func: Optional[Callable] = None):
        self.chunks: List[Dict] = []
        self.vectors: Optional[np.ndarray] = None
        self._cache: Dict[str, np.ndarray] = {}  # 文本 → 向量缓存
        self.embedding_func = embedding_func or self._default_embedding

    def _default_embedding(self, text: str) -> np.ndarray:
        """降级方案：基于关键词的哈希向量（256维）"""
        # 检查缓存
        cache_key = hashlib.md5(text.encode()).hexdigest()
        if cache_key in self._cache:
            return self._cache[cache_key]

        vec = np.zeros(256)
        words = re.findall(r"[\u4e00-\u9fff]+|[a-zA-Z_]+|\d+", text.lower())
        for word in words:
            h = int(hashlib.md5(word.encode()).hexdigest(), 16) % 256
            vec[h] += 1.0
        norm = np.linalg.norm(vec)
        vec = vec / norm if norm > 0 else vec
        self._cache[cache_key] = vec
        return vec

    def add_chunks(self, chunks: List[Dict]):
        """添加文档分块并生成向量（使用缓存）"""
        if not chunks:
            return
        self.chunks.extend(chunks)
        new_vectors = []
        for c in chunks:
            vec = self._get_or_compute_vector(c["content"])
            new_vectors.append(vec)
        new_vectors = np.array(new_vectors)
        if self.vectors is None:
            self.vectors = new_vectors
        else:
            self.vectors = np.vstack([self.vectors, new_vectors])

    def _get_or_compute_vector(self, text: str) -> np.ndarray:
        """带缓存的向量计算"""
        cache_key = hashlib.md5(text.encode()).hexdigest()
        if cache_key in self._cache:
            return self._cache[cache_key]
        vec = self.embedding_func(text)
        self._cache[cache_key] = vec
        return vec

    def search(self, query: str, top_k: int = 5) -> List[Dict]:
        """语义检索"""
        if self.vectors is None or len(self.chunks) == 0:
            return []

        query_vec = self._get_or_compute_vector(query)
        scores = self.vectors @ query_vec
        top_indices = np.argsort(scores)[-top_k:][::-1]

        results = []
        for idx in top_indices:
            if scores[idx] > 0.01:
                chunk = self.chunks[idx]
                results.append({
                    "chunk_id": chunk["chunk_id"],
                    "content": chunk["content"],
                    "score": float(scores[idx]),
                    "metadata": chunk.get("metadata", {}),
                })
        return results

    def build_context(self, query: str, top_k: int = 5, max_chars: int = 8000) -> str:
        """构建优化上下文"""
        results = self.search(query, top_k)
        parts = []
        used = 0
        for r in results:
            content = r["content"]
            meta = r["metadata"]
            header = f"\n--- 文档片段 (来源: {meta.get('source', '?')}, 相关度: {r['score']:.2f}) ---\n"
            remaining = max_chars - used - len(header)
            if remaining <= 0:
                break
            truncated = content[:remaining]
            parts.append(header + truncated)
            used += len(header) + len(truncated)
        return "".join(parts)


# ============================================================
# LLM Embedding 函数（API + 本地备选）
# ============================================================

def create_llm_embedding_func(client, model: str = "text-embedding-v1"):
    """创建基于 LLM API 的 embedding 函数，带本地备选降级"""
    _local_encoder = None

    def _get_local_encoder():
        """懒加载本地 sentence-transformers"""
        nonlocal _local_encoder
        if _local_encoder is not None:
            return _local_encoder
        try:
            from sentence_transformers import SentenceTransformer
            _local_encoder = SentenceTransformer("paraphrase-multilingual-MiniLM-L12-v2")
            return _local_encoder
        except ImportError:
            return None

    def embed(text: str) -> np.ndarray:
        # 尝试 API
        try:
            resp = client.embeddings.create(model=model, input=text[:8000])
            return np.array(resp.data[0].embedding)
        except Exception as e:
            print(f"[Warning] Embedding API 调用失败，降级到本地模型: {e}")
            # 降级到本地模型
            local_encoder = _get_local_encoder()
            if local_encoder:
                try:
                    return local_encoder.encode(text, normalize_embeddings=True)
                except Exception as e2:
                    print(f"[Warning] 本地 embedding 也失败: {e2}")

            # 最终降级：哈希向量
            vec = np.zeros(256)
            words = re.findall(r"[\u4e00-\u9fff]+|[a-zA-Z_]+|\d+", text.lower())
            for word in words:
                h = int(hashlib.md5(word.encode()).hexdigest(), 16) % 256
                vec[h] += 1.0
            norm = np.linalg.norm(vec)
            return vec / norm if norm > 0 else vec

    return embed